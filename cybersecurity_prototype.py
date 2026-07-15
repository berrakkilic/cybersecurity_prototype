#!/usr/bin/env python3
"""
Security Automation & Vulnerability Triage Prototype
Please see README.md for background, design notes and job-requirement mapping

Setup: pip install openpyxl
Run:   python3 triage_prototype.py
"""

import json
import textwrap
from datetime import datetime, timezone
from dataclasses import dataclass, field

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ----------------------------------------------------------------------------
# DATA MODEL
# ----------------------------------------------------------------------------

@dataclass
class Vulnerability:
    cve_id: str
    title: str
    description: str
    severity: str  # "Low" | "Medium" | "High" | "Critical"
    cvss_score: float
    component: str
    source: str
    discovered: str = field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


# ----------------------------------------------------------------------------
# STEP 1 — FETCH CVE DATA
# ----------------------------------------------------------------------------

def fetch_cve_feed() -> list[Vulnerability]:
    """Mocked feed. Real version: NVD API or internal scanner/inventory query."""
    return [
        Vulnerability(
            cve_id="CVE-2025-31337",
            title="Buffer Overflow in Braking ECU Firmware CAN Parser",
            description=(
                "A stack-based buffer overflow exists in the CAN message parsing "
                "routine of the embedded braking control firmware (written in C++). "
                "An attacker with access to the CAN bus can send a malformed frame "
                "with an oversized payload field, overwriting adjacent stack memory "
                "and potentially achieving arbitrary code execution on the ECU."
            ),
            severity="Critical",
            cvss_score=9.1,
            component="Braking ECU Firmware (C++, embedded, CAN bus interface)",
            source="Internal OT Vulnerability Scan / Vendor Advisory",
        ),
        Vulnerability(
            cve_id="CVE-2024-55210",
            title="Improper Certificate Validation in OTA Update Client",
            description=(
                "The over-the-air update client fails to properly validate the "
                "TLS certificate chain when downloading firmware updates, allowing "
                "a man-in-the-middle attacker on the network to serve a malicious "
                "firmware image."
            ),
            severity="High",
            cvss_score=8.1,
            component="OTA Update Client (IT/OT boundary service)",
            source="NVD",
        ),
        Vulnerability(
            cve_id="CVE-2025-01984",
            title="Cross-Site Scripting in Internal Reporting Dashboard",
            description=(
                "A stored XSS vulnerability in the internal KPI/reporting web "
                "dashboard allows an authenticated low-privilege user to inject "
                "a script that executes in the browser of an administrator "
                "viewing the report."
            ),
            severity="Medium",
            cvss_score=5.4,
            component="Internal Reporting Dashboard (React frontend)",
            source="Internal Pentest",
        ),
        Vulnerability(
            cve_id="CVE-2023-40217",
            title="Outdated OpenSSL Version with Known Handshake Bypass",
            description=(
                "A legacy diagnostics service links against an outdated OpenSSL "
                "version affected by a TLS handshake bypass, allowing session "
                "authentication to be skipped under specific conditions."
            ),
            severity="Low",
            cvss_score=3.1,
            component="Diagnostics Service",
            source="Software Composition Analysis (SCA) scan",
        ),
    ]


# ----------------------------------------------------------------------------
# STEP 2 — TRIAGE
# ----------------------------------------------------------------------------

def filter_high_priority(vulns: list[Vulnerability]) -> list[Vulnerability]:
    priority_levels = {"High", "Critical"}
    return [v for v in vulns if v.severity in priority_levels]


# ----------------------------------------------------------------------------
# STEP 3 — AI MITIGATION PLAN
# ----------------------------------------------------------------------------

def generate_mitigation_plan(vuln: Vulnerability) -> str:
    """
    Mocked LLM call, simulated deterministically via keyword matching.
    Real version: pass vuln.description to an LLM (e.g. Anthropic Messages API)
    and return the generated text.
    """
    text = vuln.description.lower()

    if "buffer overflow" in text or "stack" in text:
        steps = [
            "Patch the CAN parser to validate payload length against buffer "
            "capacity before copying (bounds checks / safe-copy primitives "
            "instead of raw memcpy).",
            "Add fuzz testing (AFL++/libFuzzer) targeting the CAN message "
            "parser to catch similar overflows before release.",
            "Enable stack canaries / ASLR / -fstack-protector-strong on the "
            "firmware build if not already active, as defense-in-depth.",
            "Restrict CAN bus access via network segmentation until patched "
            "firmware is rolled out to affected ECUs.",
        ]
    elif "certificate" in text or "tls" in text or "man-in-the-middle" in text:
        steps = [
            "Enforce full certificate chain + hostname validation in the "
            "OTA client TLS configuration; fail closed on validation errors.",
            "Pin the expected firmware-signing certificate/public key in "
            "the client to prevent acceptance of attacker-supplied certs.",
            "Add integration tests simulating an invalid/self-signed cert "
            "and assert the update is rejected.",
            "Audit other services at the IT/OT boundary for the same "
            "certificate-validation pattern.",
        ]
    else:
        steps = [
            "Reproduce the issue in staging and confirm root cause.",
            "Apply the vendor-recommended patch or upgrade the affected dependency.",
            "Add a regression test covering this vulnerability class.",
            "Re-scan the component post-fix to confirm remediation.",
        ]

    return (
            f"Summary: {vuln.title} poses a {vuln.severity.lower()} risk "
            f"(CVSS {vuln.cvss_score}) to {vuln.component}.\n"
            f"    Recommended steps:\n"
            + "\n".join(f"      {i + 1}. {s}" for i, s in enumerate(steps))
    )


# ----------------------------------------------------------------------------
# STEP 4 — JIRA TICKET PAYLOAD
# ----------------------------------------------------------------------------

def build_jira_payload(vuln: Vulnerability, mitigation_plan: str) -> dict:
    """Real version: requests.post(f"{JIRA_URL}/rest/api/3/issue", json=payload, auth=(email, token))"""
    severity_to_priority = {"Critical": "Highest", "High": "High"}

    return {
        "fields": {
            "project": {"key": "SECOPS"},
            "summary": f"[{vuln.cve_id}] {vuln.title}",
            "issuetype": {"name": "Bug"},
            "priority": {"name": severity_to_priority.get(vuln.severity, "Medium")},
            "labels": ["vulnerability", "auto-triaged", vuln.component.split()[0].lower()],
            "description": {
                "type": "doc",
                "content": [
                    {"type": "paragraph", "text": vuln.description},
                    {"type": "paragraph", "text": f"AI-Drafted Mitigation Plan:\n{mitigation_plan}"},
                ],
            },
            "customfield_cvss_score": vuln.cvss_score,
            "customfield_source": vuln.source,
            "customfield_discovered": vuln.discovered,
        }
    }


# ----------------------------------------------------------------------------
# STEP 5 — TERMINAL REPORT
# ----------------------------------------------------------------------------

def print_header(text: str):
    print("\n" + "=" * 78)
    print(f" {text}")
    print("=" * 78)


def print_report(all_vulns, priority_vulns, tickets):
    print_header("SECURITY AUTOMATION & VULNERABILITY TRIAGE PROTOTYPE")
    print(f" Run timestamp : {datetime.now(timezone.utc).isoformat()}")
    print(f" CVEs fetched   : {len(all_vulns)}")
    print(f" High/Critical  : {len(priority_vulns)}  (auto-escalated to JIRA)")

    for vuln, ticket in tickets:
        print_header(f"{vuln.cve_id} — {vuln.severity.upper()} (CVSS {vuln.cvss_score})")
        print(f" Title      : {vuln.title}")
        print(f" Component  : {vuln.component}")
        print(f" Source     : {vuln.source}")
        print(" Description:")
        print(textwrap.fill(vuln.description, width=76, initial_indent="   ", subsequent_indent="   "))

        print("\n AI-Drafted Mitigation Plan:")
        print(f"   {ticket['_mitigation_plan']}")

        print("\n Simulated JIRA Ticket Payload (POST /rest/api/3/issue):")
        print(textwrap.indent(json.dumps(ticket["_payload"], indent=2), "   "))


# ----------------------------------------------------------------------------
# STEP 6 — EXCEL DASHBOARD EXPORT
# ----------------------------------------------------------------------------

def export_to_excel(all_vulns: list[Vulnerability], tickets, filename="vulnerability_report.xlsx"):
    wb = Workbook()

    # Sheet 1: full dashboard
    ws = wb.active
    ws.title = "Vulnerability Dashboard"

    headers = ["CVE ID", "Title", "Severity", "CVSS", "Component", "Source", "Discovered"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")

    severity_fill = {
        "Critical": PatternFill("solid", fgColor="C0392B"),
        "High": PatternFill("solid", fgColor="E67E22"),
        "Medium": PatternFill("solid", fgColor="F1C40F"),
        "Low": PatternFill("solid", fgColor="7DCEA0"),
    }

    for v in all_vulns:
        ws.append([v.cve_id, v.title, v.severity, v.cvss_score, v.component, v.source, v.discovered])
        ws.cell(row=ws.max_row, column=3).fill = severity_fill.get(v.severity)

    summary_row = ws.max_row + 2
    ws.cell(row=summary_row, column=1, value="KPI Summary").font = Font(bold=True)
    for i, sev in enumerate(["Critical", "High", "Medium", "Low"]):
        count = sum(1 for v in all_vulns if v.severity == sev)
        ws.cell(row=summary_row + 1 + i, column=1, value=sev)
        ws.cell(row=summary_row + 1 + i, column=2, value=count)

    for col_idx, header in enumerate(headers, start=1):
        max_len = max(
            [len(header)] + [len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, ws.max_row + 1)]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    # Sheet 2: escalated / JIRA-bound items
    ws2 = wb.create_sheet("Escalated (JIRA)")
    headers2 = ["CVE ID", "Priority", "JIRA Project", "Summary", "AI Mitigation Plan"]
    ws2.append(headers2)
    for col in range(1, len(headers2) + 1):
        cell = ws2.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    for vuln, ticket in tickets:
        fields = ticket["_payload"]["fields"]
        ws2.append([
            vuln.cve_id,
            fields["priority"]["name"],
            fields["project"]["key"],
            fields["summary"],
            ticket["_mitigation_plan"],
        ])
        ws2.cell(row=ws2.max_row, column=5).alignment = Alignment(wrap_text=True, vertical="top")

    ws2.column_dimensions["D"].width = 45
    ws2.column_dimensions["E"].width = 70

    wb.save(filename)
    return filename


# ----------------------------------------------------------------------------
# STEP 7 — MANAGEMENT SLIDE OUTLINE
# ----------------------------------------------------------------------------

def _plain_english(vuln: Vulnerability) -> str:
    if "braking" in vuln.component.lower():
        return "A weakness was found that could let an attacker interfere with braking system software."
    if "ota" in vuln.component.lower() or "certificate" in vuln.description.lower():
        return "A weakness was found that could let an attacker push fake software updates to vehicles."
    return "A weakness was found in an internal system that needs to be fixed before it can be exploited."


def generate_slide_outline(vuln: Vulnerability, mitigation_plan: str) -> str:
    """Mocked LLM call. Real version: prompt an LLM to reframe vuln + mitigation_plan for a management audience."""
    if vuln.severity == "Critical" and "braking" in vuln.component.lower():
        business_impact = (
            "Potential remote compromise of a safety-critical braking "
            "component. Worst case: loss of vehicle safety function, "
            "regulatory exposure (CRA, product liability), and reputational "
            "damage if exploited in the field."
        )
        audience_risk = "Safety, regulatory (CRA), and brand risk — highest priority."
    elif "certificate" in vuln.description.lower() or "ota" in vuln.component.lower():
        business_impact = (
            "An attacker could deliver counterfeit firmware to vehicles via "
            "the update channel. Worst case: fleet-wide compromise through a "
            "single distribution mechanism."
        )
        audience_risk = "Fleet-scale integrity and supply-chain trust risk."
    else:
        business_impact = (
            "Could allow unauthorized access or disruption to internal "
            "systems if left unaddressed."
        )
        audience_risk = "Operational risk to internal tooling."

    return f"""\
## Slide 1: Title
- Vulnerability Briefing: {vuln.cve_id}
- Prepared for: Management Alignment Meeting
- Date: {datetime.now(timezone.utc).strftime('%Y-%m-%d')}

## Slide 2: Executive Summary
- **What happened:** A {vuln.severity.lower()}-severity vulnerability was
  identified in {vuln.component}.
- **Business impact:** {business_impact}
- **Risk category:** {audience_risk}
- **Status:** Ticket auto-created in JIRA ({vuln.cve_id}), mitigation in progress.

## Slide 3: Why This Matters (Non-Technical)
- Plain-language translation of the issue — avoid CVE/CVSS jargon here.
- Frame in terms of: safety, cost, compliance (CRA), customer trust.
- One sentence, no acronyms: "{_plain_english(vuln)}"

## Slide 4: Technical Detail (Backup / Appendix)
- CVE ID: {vuln.cve_id} | CVSS: {vuln.cvss_score} | Severity: {vuln.severity}
- Component: {vuln.component}
- Source: {vuln.source}
- Full description available in JIRA ticket (link).

## Slide 5: Mitigation Plan & Timeline
- AI-drafted mitigation plan (developer-reviewed):
{textwrap.indent(mitigation_plan, "  ")}
- Owner: [Assign engineering owner]
- Target resolution: [Insert SLA based on severity, e.g. Critical = 72h]

## Slide 6: Ask / Decision Needed
- Decision requested: approve resourcing / prioritization for fix.
- Any budget, schedule, or cross-team dependency to flag.
- Next status update: [Insert date]
"""


# ----------------------------------------------------------------------------
# MAIN
# ----------------------------------------------------------------------------

def main():
    all_vulns = fetch_cve_feed()
    priority_vulns = filter_high_priority(all_vulns)

    tickets = []
    for vuln in priority_vulns:
        mitigation_plan = generate_mitigation_plan(vuln)
        payload = build_jira_payload(vuln, mitigation_plan)
        tickets.append((vuln, {"_payload": payload, "_mitigation_plan": mitigation_plan}))

    print_report(all_vulns, priority_vulns, tickets)

    print_header("STEP 6: EXCEL DASHBOARD EXPORT")
    xlsx_path = export_to_excel(all_vulns, tickets)
    print(f" Excel report written to: {xlsx_path}")

    most_critical = max(priority_vulns, key=lambda v: v.cvss_score)
    most_critical_plan = next(t["_mitigation_plan"] for v, t in tickets if v.cve_id == most_critical.cve_id)
    slide_outline = generate_slide_outline(most_critical, most_critical_plan)

    print_header(f"STEP 7: MANAGEMENT SLIDE OUTLINE — {most_critical.cve_id}")
    print(slide_outline)

    outline_path = "management_slide_outline.md"
    with open(outline_path, "w") as f:
        f.write(slide_outline)
    print(f" Markdown slide outline saved to: {outline_path}")

    print_header("RUN COMPLETE")
    print(f" {len(tickets)} JIRA ticket(s) simulated. See README.md for production path.\n")


if __name__ == "__main__":
    main()